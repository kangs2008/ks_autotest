from openpyxl import load_workbook
from openpyxl.styles import Font
import re, os, sys
from pathlib import Path

class Handle_excel():
    def __init__(self, filepath):
        self.wb = load_workbook(filepath)

    def get_sheets(self):
        sheet_names = self.wb.sheetnames
        return sheet_names

    def get_sheets_by_rule(self, rule):
        sheets = self.get_sheets()
        slist = []
        for sheet in sheets:
            if sheet.startswith(rule):
                slist.append(sheet)
        if len(slist) == 0:
            print(f'Input sheet name rule "{rule}" not find in excel, Excel sheets "{sheets}"')
        return slist

    def get_sheet_by_name(self, sheet_name):
        try:
            sheet = self.wb[sheet_name]
        except:
            names = self.get_sheets()
            print(f'Input sheet name "{sheet_name}" not exist in excel, please check it. Excel sheets {names}')
            sheet = None
        return sheet
    def get_max_row(self, sheet):
        max_row = sheet.max_row
        return max_row
    def get_max_column(self, sheet):
        max_col = sheet.max_column
        return max_col
    def get_row_value(self, sheet, row):
        columns = sheet.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = sheet.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def get_cell_value(self, sheet, row, col):
        value = sheet.cell(row, col).value
        return value
    def _get_re_parameter(self, dict_kv, para):
        pattern = r'[$][{](.*?)[}]'
        para = str(para)
        for key in dict_kv.keys():
            res = re.findall(pattern, str(para)) # ${aa}
            if res:
                for r in res:
                    if r == key:
                        para = para.replace('${' + r + '}', str(dict_kv.get(r, r)))
                # para = [para.replace('{' + r + '}', str(dict_kv.get(r, r))) for r in enumerate(res) if r == key]
        return para

    def get_column_values_by_title(self, sheet_obj, exec_value):
        """
        get_column_values_by_title
        :param sheet_obj: worksheet obj
        :param exec_value: row title 'exec'
        :return: column_values / column pos
        """
        row_num = self.get_max_row(sheet_obj)
        title_row = self.get_row_value(sheet_obj, 1)
        column_values_list = []
        if exec_value in title_row:
            pos = [i for i, title in enumerate(title_row) if title == exec_value]
            openpyxl_pos = pos[0] + 1  # 1 start
            for rownum in range(2, row_num + 1):
                # value = sheet_obj.cell(rownum, openpyxl_pos).value
                value = self.get_cell_value(sheet_obj, rownum, openpyxl_pos)
                if value is None:
                    value = ''
                column_values_list.append(value)
            return column_values_list, openpyxl_pos
        else:
            print(f'Input title "{exec_value}" is not in Excel "{sheet_obj}" title_row list.')

    def get_dictConfig_from_sheet(self, sheet_config):
        """
        get_dictConfig_from_sheet  {'user': 'kou', 'pwd': 123456, 'trade': 410}
        :param sheet_config: worksheet obj
        :return: all values of sheet
        """
        row_num = self.get_max_row(sheet_config)
        tempdict = {}
        for row in range(2, row_num + 1):
            # key = sheet_config.cell(row, 1).vlaue  # title key
            key = self.get_cell_value(sheet_config, row, 1)
            # value = sheet_config.cell(row, 2).value
            value = self.get_cell_value(sheet_config, row, 2)
            if value is None:
                value = ''
            tempdict[key] = value
        return tempdict

    def get_exec_dictConfig_from_sheet(self, sheet_obj_config, exec_value='exec', exec_type='y'):
        """
        get_exec_dict_from_sheet  {'pwd': 134455, 'trade': 410}
        :param sheet_obj_config: worksheet obj
        :param exec_value: get 'exec' column
        :param exec_type: get 'exec' column and cell value is 'y'
        :return: all values of sheet / all values of this column contains cell 'y'
        """
        row_num = self.get_max_row(sheet_obj_config)
        tempdict = {}
        if exec_value is None or exec_value == '':
            tempdict = self.get_dictConfig_from_sheet(sheet_obj_config)
        else:
            print(sheet_obj_config)
            # title_value = sheet_obj_config.cell(1, 3).vlaue
            title_value = self.get_cell_value(sheet_obj_config, 1, 3)
            print(title_value)
            for row in range(2, row_num + 1):
                # value = sheet_obj_config.cell(row, 3).vlaue
                value = self.get_cell_value(sheet_obj_config, row, 3)
                if value is None:
                    value = ''
                if title_value == exec_value and value.lower() == exec_type:
                    # key = sheet_obj_config.cell(row, 1).value
                    key = self.get_cell_value(sheet_obj_config, row, 1)
                    # value = sheet_obj_config.cell(row, 2).value
                    value = self.get_cell_value(sheet_obj_config, row, 2)
                    tempdict[key] = value
        return tempdict

    def get_dictList_from_sheet(self, sheet_obj):
        """
        get_dictList_from_sheet
        :param sheet_obj: worksheet obj
        :return: all values of sheet
        """
        row_num = self.get_max_row(sheet_obj)
        column_num = self.get_max_column(sheet_obj)
        sheet_values_list = []
        for row in range(2, row_num + 1):
            tempdict = {}
            for column in range(1, column_num + 1):
                # title_value = sheet_obj.cell(1, column).value
                title_value = self.get_cell_value(sheet_obj, 1, column)
                # value = sheet_obj.cell(row, column).value
                value = self.get_cell_value(sheet_obj, row, column)
                if value is None:
                    value = ''
                tempdict[title_value] = value
            sheet_values_list.append(tempdict)
        return sheet_values_list


    def get_exec_dictvList_from_sheet(self, sheet_obj, exec_value='exec', exec_type='y'):
        """
        get_exec_dictvList_from_sheet
        :param sheet_obj_config: worksheet obj
        :param exec_value: get 'exec' column
        :param exec_type: get 'exec' column and cell value is 'y'
        :return: all values of sheet / all values of this column contains cell 'y'
        """
        row_num = self.get_max_row(sheet_obj)
        column_num = self.get_max_column(sheet_obj)

        sheet_values_list = []
        if exec_value is None:
            sheet_values_list = self.get_dictList_from_sheet(sheet_obj)
            return sheet_values_list
        else:
            for row in range(2, row_num + 1):
                tempdict = {}
                for column in range(1, column_num + 1):
                    # title_value = sheet_obj.cell(1, column) # title key
                    title_value = self.get_cell_value(sheet_obj, 1, column)
                    # value = sheet_obj.cell(row, column).value
                    value = self.get_cell_value(sheet_obj, row, column)
                    if value is None:
                        value = ''
                    if title_value == exec_value and value.lower() == exec_type:
                        value = str(row - 1)
                    tempdict[title_value] = value
                sheet_values_list.append(tempdict)
        return sheet_values_list

    def get_exec_dictList_from_sheet_re(self, sheet_obj, config_list, exec_value='exec', exec_type='y'):
        """
        get_exec_dictList_from_sheet_re
        :param sheet_obj_config: worksheet obj
        :param config_list: kv config sheet list
        :param exec_value: get 'exec' column
        :param exec_type: get 'exec' column and cell value is 'y'
        :return: all values of sheet / all values of this column contains cell 'y'
        """
        row_num = self.get_max_row(sheet_obj)
        column_num = self.get_max_column(sheet_obj)
        exec_column_values, col_pos = self.get_column_values_by_title(sheet_obj, exec_value)
        _dict_conf = {}
        if isinstance(config_list, list):
            for one in config_list:
                if one != '':
                    _sheet_obj = self.get_sheet_by_name(one)
                    print(_sheet_obj)
                    _dict = self.get_exec_dictConfig_from_sheet(_sheet_obj, exec_value, exec_type)
                    _dict_conf.update(_dict)
        elif config_list != '' and config_list is not None:
            _sheet_obj = self.get_sheet_by_name(config_list)
            _dict = self.get_exec_dictConfig_from_sheet(_sheet_obj, exec_value, exec_type)
            _dict_conf.update(_dict)
        else:
            print(f'Input configList is "{config_list}".')

        sheet_values_list = []
        if exec_value is None or exec_value == '':
            sheet_values_list = self.get_dictList_from_sheet(sheet_obj)
            return sheet_values_list
        else:
            for row in range(2, row_num + 1):
                tempdict = {}
                if exec_column_values[row-2].lower() == exec_type:
                    for column in range(1, column_num + 1):
                        # title_value = sheet_obj.cell(1, column).value # title key
                        title_value = self.get_cell_value(sheet_obj, 1, column)
                        # value = sheet_obj.cell(row, column).value
                        value = self.get_cell_value(sheet_obj, row, column)
                        if value is None:
                            value = ''
                        if title_value == exec_value:
                            value = str(row-1)
                            tempdict[title_value] = value
                        change_value = self._get_re_parameter(_dict_conf, value)
                        tempdict[title_value] = change_value
                    sheet_values_list.append(tempdict)
        return sheet_values_list


def load_excel(file_name, sheet_name):
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    return wb, sheet

def write_to_excel(sheet_obj, set_value, row_pos, col_pos):
    # 字体格式，颜色和大小
    sheet_obj.cell(int(row_pos) + 1, int(col_pos)).value = set_value
    if set_value == 'FAIL':
        font_s = Font(bold=False, color="FF0000", size=10)
        sheet_obj.cell(int(row_pos) + 1, int(col_pos)).font = font_s
    elif set_value == 'PASS':
        font_s = Font(bold=False, color="00FF00", size=10)
        sheet_obj.cell(int(row_pos) + 1, int(col_pos)).font = font_s
    else:
        pass

def excel_to_save(wb, file_name):
    wb.save(file_name)

def _check_input_sheet_name_in_excel(multi_excel_list, sheet_name_list):
    _list = []
    err = []
    if isinstance(multi_excel_list, list):
        for i, file_name in enumerate(multi_excel_list):
            count = 0
            all_sheet_names = Handle_excel(Path(file_name)).get_sheets()
            if isinstance(sheet_name_list, list):
                for one_sheet in sheet_name_list:
                    if one_sheet != '':
                        if one_sheet not in all_sheet_names:
                            count += 1
                            msg = f'Input sheet name "{one_sheet}" not in excel file "{Path(file_name)}", please check it. Excel file sheets:{all_sheet_names}'
                            err.append(msg)
                    if count ==len(sheet_name_list):
                        _list.append(count)
            else:
                if sheet_name_list != '' and sheet_name_list != None:
                    if sheet_name_list not in all_sheet_names:
                        count += 1
                        msg = f'Input sheet name "{sheet_name_list}" not in excel file "{Path(file_name)}", please check it. Excel file sheets:{all_sheet_names}'
                        err.append(msg)
                    if count == 1:
                        _list.append(count)
        if _list:
            print(err)
            raise
    else:
        all_sheet_names = Handle_excel(Path(multi_excel_list)).get_sheets()
        if isinstance(sheet_name_list, list):
            count = 0
            for one_sheet in sheet_name_list:
                if one_sheet != '':
                    if one_sheet not in all_sheet_names:
                        count += 1
                        msg = f'Input sheet name "{sheet_name_list}" not in excel file "{Path(multi_excel_list)}", please check it. Excel file sheets:{all_sheet_names}'
                        err.append(msg)
            if err:
                print(err)
                raise
        else:
            count = 0
            if sheet_name_list != '' and sheet_name_list != None:
                if sheet_name_list not in all_sheet_names:
                    count += 1
                    msg = f'Input sheet name "{sheet_name_list}" not in excel file "{Path(multi_excel_list)}", please check it. Excel file sheets:{all_sheet_names}'
                    err.append(msg)
                if err:
                    print(err)
                    raise

def _check_input_sheet_name_rule_in_excel(sheet_name_list, sheet_name_rule):
    if isinstance(sheet_name_list, list):
        count = 0
        if len(sheet_name_list) == 1 and sheet_name_list[0] == '':
            pass
        else:
            for one_sheet in sheet_name_list:
                if isinstance(sheet_name_rule, list):
                        msg = f'Input sheet_name_rule type "string", type "list" incorrect.'
                        print(msg)
                        raise
                else:
                    if sheet_name_rule in one_sheet:
                        count += 1
            if count != len(sheet_name_list):
                print(f'Input sheet_name_rule "{sheet_name_rule}" not in sheet_name_list "{sheet_name_list}".')
                raise
    else:
        if sheet_name_list != '' and sheet_name_list != None:
            if isinstance(sheet_name_rule, list):
                    msg = f'Input sheet_name_rule type "string", type "list" incorrect.'
                    print(msg)
                    raise
            else:
                if sheet_name_rule not in sheet_name_list:
                    print(f'Input sheet_name_rule "{sheet_name_rule}" not in sheet_name_list "{sheet_name_list}".')
                    raise



def excel_to_case(multi_excel_list, sheet_name_list=[], sheet_name_rule='t_', config_list=[], exec_value='exec', exec_type='y'):
    """
    excel_to_case path / path, [] , 't_' / path , ['t_xxx.xlsx'], 't_', ['config']
    :param multi_excel_list: excel file list / one excel file
    :param sheeet_name_list: excel sheet name list: []/[''] is all 't_' start sheets, ['xxx.elsx']
    :param sheet_name_rule: get excel sheet name by default rule 't_'
    :param config_list: kv config sheet list
    :param exec_value: get 'exec' column
    :param exec_type: get 'exec' column and cell value is 'y'
    :return: all values of sheet
    """
    _all_values = []
    if isinstance(multi_excel_list, list):
        for i, file_name in enumerate(multi_excel_list):
            _check_input_sheet_name_in_excel(multi_excel_list, sheet_name_list)
            _check_input_sheet_name_in_excel(multi_excel_list, config_list)
            _check_input_sheet_name_rule_in_excel(sheet_name_list, sheet_name_rule)
            name = Path(file_name).name
            if isinstance(sheet_name_list, list):
                if len(sheet_name_list) != 0 and sheet_name_list[0] != '':
                    name_rule_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                    for i, sheet_one in enumerate(sheet_name_list):
                        if len(name_rule_list) != 0:
                            if sheet_one in name_rule_list:
                                case_kv = {}
                                sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_one)
                                excel_kv_values = Handle_excel(file_name).get_exec_dictList_from_sheet_re(sheet_obj, config_list, exec_value, exec_type)
                                case_kv[f'sheetname'] = sheet_one
                                case_kv[f'file'] = name
                                case_kv[f'filepath'] = Path(file_name)
                                case_kv[f'filesheet'] = excel_kv_values
                                _all_values.append(case_kv)
                else:
                    name_rule_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                    for i, sheet_one in enumerate(name_rule_list):
                        case_kv = {}
                        sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_one)
                        excel_kv_values = Handle_excel(file_name).get_exec_dictList_from_sheet_re(sheet_obj, config_list, exec_value, exec_type)
                        case_kv[f'sheetname'] = sheet_one
                        case_kv[f'file'] = name
                        case_kv[f'filepath'] = Path(file_name)
                        case_kv[f'filesheet'] = excel_kv_values
                        _all_values.append(case_kv)
            else:
                name_rule_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                if sheet_name_list in name_rule_list:
                    case_kv = {}
                    sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_name_list)
                    excel_kv_values = Handle_excel(file_name).get_exec_dictList_from_sheet_re(sheet_obj, config_list,
                                                                                            exec_value, exec_type)
                    case_kv[f'sheetname'] = sheet_name_list
                    case_kv[f'file'] = name
                    case_kv[f'filepath'] = Path(file_name)
                    case_kv[f'filesheet'] = excel_kv_values
                    _all_values.append(case_kv)
    else:
        _name = multi_excel_list
        file_name = Path(_name)
        _check_input_sheet_name_in_excel(file_name, sheet_name_list)
        _check_input_sheet_name_in_excel(file_name, config_list)
        _check_input_sheet_name_rule_in_excel(sheet_name_list, sheet_name_rule)
        name = Path(file_name).name
        if isinstance(sheet_name_list, list):
            if len(sheet_name_list) != 0 and sheet_name_list[0] != '':
                name_rule_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                for i, sheet_one in enumerate(sheet_name_list):
                    if len(name_rule_list) != 0:
                        if sheet_one in name_rule_list:
                            case_kv = {}
                            sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_one)
                            excel_kv_values = Handle_excel(file_name).get_exec_dictList_from_sheet_re(sheet_obj,
                                                                                                      config_list,
                                                                                                      exec_value, exec_type)
                            case_kv[f'sheetname'] = sheet_one
                            case_kv[f'file'] = name
                            case_kv[f'filepath'] = Path(file_name)
                            case_kv[f'filesheet'] = excel_kv_values
                            _all_values.append(case_kv)
            else:
                name_rule_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                for i, sheet_one in enumerate(name_rule_list):
                    case_kv = {}
                    sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_one)
                    excel_kv_values = Handle_excel(file_name).get_exec_dictList_from_sheet_re(sheet_obj,
                                                                                              config_list,
                                                                                              exec_value, exec_type)
                    case_kv[f'sheetname'] = sheet_one
                    case_kv[f'file'] = name
                    case_kv[f'filepath'] = Path(file_name)
                    case_kv[f'filesheet'] = excel_kv_values
                    _all_values.append(case_kv)

        else:
            case_kv = {}
            sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_name_list)
            if sheet_obj is not None:
                print(config_list)
                excel_kv_values = Handle_excel(file_name).get_exec_dictList_from_sheet_re(sheet_obj, config_list,
                                                                                        exec_value, exec_type)
                case_kv[f'sheetname'] = sheet_name_list
                case_kv[f'file'] = name
                case_kv[f'filepath'] = Path(file_name)
                case_kv[f'filesheet'] = excel_kv_values
                _all_values.append(case_kv)
    return _all_values


if __name__ == '__main__':
    pass
