from openpyxl import load_workbook
from openpyxl.styles import Font
import re, os, sys

class Handle_excel():
    def __init__(self, filepath):
        self.wb = load_workbook(filepath)

    def get_sheets(self):
        sheet_names = self.wb.sheetnames
        return sheet_names

    def get_sheets_by_rule(self, rule='t_'):
        sheets = self.get_sheets()
        slist = []
        for sheet in sheets:
            if sheet.startswith(rule):
                slist.append(sheet)
        return slist

    def get_sheet_by_name(self, sheet_name):
        try:
            sheet = self.wb[sheet_name]
        except:
            names = self.get_sheets()
            print(f'Input sheet name "{sheet_name}" not exist in excel, please check it. Excel sheets {names}')
            sheet = None
        return sheet
    def get_max_row(self, sheet):
        max_row = sheet.max_row
        return max_row
    def get_max_column(self, sheet):
        max_col = sheet.max_column
        return max_col
    def get_row_value(self, sheet, row):
        columns = sheet.max_column
        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = sheet.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def get_cell_value(self, sheet, row, col):
        value = sheet.cell(row, col).value
        return value
    def get_re_parameter(self, dict_kv, para):
        pattern = r'[$][{](.*?)[}]'
        para = str(para)
        for key in dict_kv.keys():
            res = re.findall(pattern, str(para)) # ${aa}
            if res:
                for r in res:
                    if r == key:
                        para = para.replace('${' + r + '}', str(dict_kv.get(r, r)))
                # para = [para.replace('{' + r + '}', str(dict_kv.get(r, r))) for r in enumerate(res) if r == key]
        return para

    def get_column_values_by_title(self, sheet, exec_value):
        maxRowNum = self.get_max_row(sheet)
        titleRow = self.get_row_value(sheet, 1)
        execValues = []
        if exec_value in titleRow:
            pos = [i for i, title in enumerate(titleRow) if title == exec_value]

            if len(pos) != 0:
                for rownum in range(2, maxRowNum + 1):
                    value = sheet.cell(rownum, pos[0] + 1).value
                    if value is None:
                        value = ''
                    execValues.append(value)

                return execValues, pos
            else:
                print("'exec' is nt define")
        else:
            print('the input exec_value is not in titleRow list.')

    def get_kvList_from_sheet(self, sheet):
        maxRowNum = self.get_max_row(sheet)
        columnNum = self.get_max_column(sheet)

        sheetValues = []
        for row in range(2, maxRowNum + 1):
            tempdict = {}
            for column in range(1, columnNum + 1):
                title_value = self.get_cell_value(sheet, 1, column) # title key
                value = sheet.cell(row, column).value

                if value is None:
                    value = ''
                tempdict[title_value] = value
                sheetValues.append(tempdict)
        return sheetValues


    def get_exec_kvList_from_sheet(self, sheet, exec_value='exec', exec_type='y'):
        maxRowNum = self.get_max_row(sheet)
        columnNum = self.get_max_column(sheet)

        sheetValues = []
        if exec_value is None:
            sheetValues = self.get_kvList_from_sheet(sheet)
            return sheetValues
        else:
            exec, pos = self.get_column_values_by_title(sheet, exec_value)
            for row in range(2, maxRowNum + 1):

                execvalue = exec[row - 2]
                real_pos = pos[0] + 1
                tempdict = {}
                if execvalue == exec_type.lower():
                    for column in range(1, columnNum + 1):
                        title_value = self.get_cell_value(sheet, 1, column) # title key
                        value = sheet.cell(row, column).value

                        if value is None:
                            value = ''

                        if column == real_pos:
                            tempdict[title_value] = str(row - 1) # 1 2 3
                    sheetValues.append(tempdict)
        return sheetValues

    def get_dict_from_sheet(self, sheet_config):
        maxRowNum = self.get_max_row(sheet_config)
        tempdict = {}
        for row in range(2, maxRowNum + 1):
            key = sheet_config.cell(row, 1).value
            value = sheet_config.cell(row, 1).value
            if value is None:
                value = ''
            tempdict[key] = value
        return tempdict


    def get_exec_dict_from_sheet(self, sheet_config, exec_value='exec', exec_type='y'):
        maxRowNum = self.get_max_row(sheet_config)

        tempdict = {}
        if exec_value is None or exec_value == '':
            tempdict = self.get_dict_from_sheet(sheet_config)

        else:
            exec_value_list, pos = self.get_column_values_by_title(sheet_config, exec_value)
            for row in range(2, maxRowNum + 1):
                execvalue = exec_value_list[row - 2]
                if execvalue == exec_type.lower():
                    key = sheet_config.cell(row, 1).value
                    value = sheet_config.cell(row, 2).value
                    if value is None:
                        value = ''
                    tempdict[key] = value
        return tempdict

    def get_exec_kvList_from_sheet_re(self, sheet, multiList=[], exec_value='exec', exec_type='y'):
        maxRowNum = self.get_max_row(sheet)
        columnNum = self.get_max_column(sheet)

        sheetValues = []
        if exec_value is None:
            sheetValues = self.get_kvList_from_sheet(sheet)
            return sheetValues
        else:
            exec, pos = self.get_column_values_by_title(sheet, exec_value)
            for row in range(2, maxRowNum + 1):

                execvalue = exec[row - 2]
                real_pos = pos[0] + 1
                tempdict = {}
                if execvalue.lower() == exec_type.lower():
                    for column in range(1, columnNum + 1):
                        title_value = self.get_cell_value(sheet, 1, column) # title key
                        value = sheet.cell(row, column).value

                        if value is None:
                            value = ''
                            tempdict[title_value] = value
                        else:
                            if isinstance(multiList, list):
                                if len(multiList) != 0:

                                    for one in multiList:
                                        if one == '':
                                            tempdict[title_value] = value
                                        else:
                                            _sheet = self.get_sheet_by_name(one)
                                            _dict = self.get_exec_dict_from_sheet(_sheet, exec_value, exec_type)
                                            change_value = self.get_re_parameter(_dict, value)
                                            tempdict[title_value] = change_value
                                            value = change_value
                                else:
                                    tempdict[title_value] = value
                            else:
                                if multiList != '' or multiList != None:
                                    _sheet = self.get_sheet_by_name(multiList)
                                    _dict = self.get_exec_dict_from_sheet(_sheet, exec_value, exec_type)
                                    change_value = self.get_re_parameter(_dict, value)
                                    tempdict[title_value] = change_value
                                else:
                                    tempdict[title_value] = value

                        if column == real_pos:
                            tempdict[title_value] = str(row - 1) # 1 2 3
                    sheetValues.append(tempdict)
        return sheetValues
    #
    # def get_re_parameter_sheet(self, dict_kv, para):
    #     pattern = r'[$][{](.*?)[}]'
    #     para = str(para)
    #     for val in dict_kv.values():
    #         res = re.findall(pattern, str(para)) # ${aa}
    #         if res:
    #             for r in res:
    #                 if r in para:
    #                     para = para.replace('${' + r + '}', str(val))
    #     return para



def load_excel(file_name, sheet_name):
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    return wb, sheet

def write_to_excel(sheet, set_value, row_pos, col_pos):
    # 字体格式，颜色和大小
    sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).value = set_value
    if set_value == 'FAIL':
        font_s = Font(bold=False, color="FF0000", size=10)
        sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).font = font_s
    elif set_value == 'PASS':
        font_s = Font(bold=False, color="00FF00", size=10)
        sheet.cell(int(row_pos) + 1, int(col_pos[0]) + 1).font = font_s
    else:
        pass

def excel_to_save(wb, file_name):
    wb.save(file_name)

def get_file_all_dir(file_dir):
    L = []
    L_name = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if file.endswith('.xlsx') and file.startswith('test_'):
                f = os.path.join(root, file)
                fnew = f.replace('\\', '/')
                L.append(fnew)
                L_name.append(file)
    return L, L_name

def get_file_current_dir(path):
    L = []
    L_name = []
    for file in os.listdir(path):
        if file.endswith('.xlsx') and file.startswith('test_'):
            f = os.path.join(path, file)
            fnew = f.replace('\\', '/')
            L.append(fnew)
            L_name.append(file)
    return L, L_name

def check_input_sheet_name_in_excel(file_name, multiList=[]):
    all_sheet_names = Handle_excel(file_name).get_sheets()
    count = 0
    if isinstance(multiList, list):
        for one_sheet in multiList:
            if one_sheet == '':
                pass
            else:
                if one_sheet not in all_sheet_names:
                    print(
                        f'Input sheet name "{multiList}" not in excel file "{file_name}", please check it. Excel file sheets:{all_sheet_names}')
                    raise
    else:
        if multiList == '' or multiList == None:
            pass
        else:
            if multiList not in all_sheet_names:
                print(
                    f'Input sheet name "{multiList}" not in excel file "{file_name}", please check it. Excel file sheets:{all_sheet_names}')
                raise

def excel_to_case_from_file(multifile, sheeet_name=[], multiList=[], sheet_name_rule='t_', exec_value='exec', exec_type='y'):

    _all_values = []

    if isinstance(multifile, list):

        for i, file_name in enumerate(multifile):
            dir, name = os.path.split(file_name)
            if sys.platform != 'win32':
                pass
            else:
                file_name = file_name.replace('\\', '/')
            check_input_sheet_name_in_excel(file_name, sheeet_name)
            check_input_sheet_name_in_excel(file_name, multiList)
            if isinstance(sheeet_name, list):

                if len(sheeet_name) != 0 and sheeet_name[0] != '':
                    sheet_name_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                    for i, sheet_one in enumerate(sheeet_name):
                        # if sheet_one not in sheet_name_list:
                        #     print(f'Input sheet name "{sheet_one}" not in excel file, please check it, Excel file sheets:{sheet_name_list}')
                        #     raise
                        case_kv = {}
                        sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_one)
                        excel_kv_values = Handle_excel(file_name).get_exec_kvList_from_sheet_re(sheet_obj, multiList, exec_value, exec_type)
                        case_kv[f'sheetname'] = sheet_one
                        case_kv[f'file'] = name
                        case_kv[f'filepath'] = file_name
                        case_kv[f'filesheet'] = excel_kv_values
                        _all_values.append(case_kv)
                else:
                    sheet_name_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                    for i, sheet_one in enumerate(sheet_name_list):

                        case_kv = {}
                        sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_one)
                        excel_kv_values = Handle_excel(file_name).get_exec_kvList_from_sheet_re(sheet_obj, multiList, exec_value, exec_type)
                        case_kv[f'sheetname'] = sheet_one
                        case_kv[f'file'] = name
                        case_kv[f'filepath'] = file_name
                        case_kv[f'filesheet'] = excel_kv_values
                        _all_values.append(case_kv)
            else:
                sheet_name_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                # if sheeet_name not in sheet_name_list:
                    # print(
                    #     f'Input sheet name "{sheeet_name}" not in excel file, please check it, Excel file sheets:{sheet_name_list}')
                    # raise
                case_kv = {}
                sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheeet_name)
                excel_kv_values = Handle_excel(file_name).get_exec_kvList_from_sheet_re(sheet_obj, multiList,
                                                                                        exec_value, exec_type)
                case_kv[f'sheetname'] = sheeet_name
                case_kv[f'file'] = name
                case_kv[f'filepath'] = file_name
                case_kv[f'filesheet'] = excel_kv_values
                _all_values.append(case_kv)
    else:
        file_name = multifile
        dir, name = os.path.split(file_name)
        if sys.platform != 'win32':
            pass
        else:
            file_name = file_name.replace('\\', '/')
        check_input_sheet_name_in_excel(file_name, sheeet_name)
        check_input_sheet_name_in_excel(file_name, multiList)
        if isinstance(sheeet_name, list):
            if len(sheeet_name) != 0 and sheeet_name[0] != '':
                sheet_name_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                for i, sheet_one in enumerate(sheeet_name):
                    # if sheet_one not in sheet_name_list:
                    #     print(f'Input sheet name "{sheet_one}" not in excel file, please check it, Excel file sheets:{sheet_name_list}')
                    #     raise
                    case_kv = {}
                    sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_one)
                    excel_kv_values = Handle_excel(file_name).get_exec_kvList_from_sheet_re(sheet_obj, multiList, exec_value, exec_type)
                    case_kv[f'sheetname'] = sheet_one
                    case_kv[f'file'] = name
                    case_kv[f'filepath'] = file_name
                    case_kv[f'filesheet'] = excel_kv_values
                    _all_values.append(case_kv)
            else:
                sheet_name_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
                for i, sheet_one in enumerate(sheet_name_list):

                    case_kv = {}
                    sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheet_one)
                    excel_kv_values = Handle_excel(file_name).get_exec_kvList_from_sheet_re(sheet_obj, multiList, exec_value, exec_type)
                    case_kv[f'sheetname'] = sheet_one
                    case_kv[f'file'] = name
                    case_kv[f'filepath'] = file_name
                    case_kv[f'filesheet'] = excel_kv_values
                    _all_values.append(case_kv)
        else:
            sheet_name_list = Handle_excel(file_name).get_sheets_by_rule(sheet_name_rule)
            # if sheeet_name not in sheet_name_list:
            #     print(
            #         f'Input sheet name "{sheeet_name}" not in excel file, please check it, Excel file sheets:{sheet_name_list}')
            #     raise
            case_kv = {}
            sheet_obj = Handle_excel(file_name).get_sheet_by_name(sheeet_name)
            excel_kv_values = Handle_excel(file_name).get_exec_kvList_from_sheet_re(sheet_obj, multiList,
                                                                                    exec_value, exec_type)
            case_kv[f'sheetname'] = sheeet_name
            case_kv[f'file'] = name
            case_kv[f'filepath'] = file_name
            case_kv[f'filesheet'] = excel_kv_values
            _all_values.append(case_kv)


    return _all_values


if __name__ == '__main__':
    pass
